## Importando as bibliotecas necessárias
import requests # Flask
import json # Glean
from typing import List, Dict # Glean
from flask import Flask, request # Flask
import warnings #retira um warning de SSL do prompt
import datetime
import threading

warnings.filterwarnings("ignore", category=UserWarning, module="urllib3")
import os
from openpyxl import Workbook, load_workbook

##--------------------------------------------------------------------------##
## Setup das configurações globais e inicialização do Flask
app = Flask(__name__)
from config import *  # Importa as variáveis de ambiente do arquivo config.py

# Headers para Glean e Zendesk
GLEAN_HEADERS = {
    'Content-Type': 'application/json',
    'Authorization': f'Bearer {GLEAN_TOKEN}'
}
ZENDESK_HEADERS = {
    'Content-Type': 'application/json'
}
##--------------------------------------------------------------------------##
##--------------------------------------------------------------------------##
def buscar_dados_completos_do_ticket(ticket_id): # Busca os dados completos do ticket via API, usando do ticket ID
    url = f"https://{ZENDESK_DOMAIN}.zendesk.com/api/v2/tickets/{ticket_id}.json" # URL da API do Zendesk para pegar o ticket
    auth = (ZENDESK_EMAIL, ZENDESK_TOKEN) # Autenticação com email e token
    response = requests.get(url, auth=auth) # Faz a requisição GET com autenticação
    if response.status_code == 200: # Verifica se a requisição foi bem sucedida
        return response.json()["ticket"] # Retorna os dados do ticket
        
    else: # Se a requisição não foi bem sucedida
        print(f"Erro ao buscar ticket: {response.status_code}") # Imprime o erro
        return {} # Retorna um dicionário vazio
##--------------------------------------------------------------------------##
def buscar_comentarios_do_ticket(ticket_id): # Busca os comentários do ticket via API, usando do ticket ID
    url = f"https://{ZENDESK_DOMAIN}.zendesk.com/api/v2/tickets/{ticket_id}/comments.json" # URL da API do Zendesk para pegar os comentários do ticket
    auth = (ZENDESK_EMAIL, ZENDESK_TOKEN) # Autenticação com email e token
    response = requests.get(url, headers=ZENDESK_HEADERS, auth=auth) #Extrai os comentários dos tickets
    if response.status_code == 200:  #a resposta sendo positiva, retornamos
        return response.json().get("comments", [])
    else: #retorna vazio caso exista erro
        print(f"Erro ao buscar comentários: {response.status_code}") #printa no terminal
        return []
##--------------------------------------------------------------------------##
def gerar_texto_completo_do_ticket(ticket_id, ticket, comentarios): #pega as informaçoes de id e de comentários e gera um texto para a glean
    subject = ticket.get("subject", "Sem assunto") # Inicializa assunto do ticket
    conteudo = f"-------------\nTicket ID: {ticket_id}\n" # Escreve ID do ticket
    conteudo += f" - subject: {subject}\n" # Escreve Assunto do ticket

    for idx, comentario in enumerate(comentarios, start=1):# Itera pelos comentários
        corpo = comentario.get("body", "").replace("\n", " ").strip()
        autor_id = comentario.get("author_id")
        autor_email, grupos = get_user_info(autor_id)

        # Pula comentários de sistema@vtex.com.br ou glean@vtex.com.br
        if autor_email in ["sistema@vtex.com.br", "glean@vtex.com"]: # Verifica se o autor é um sistema ou Glean
            continue # Pula o comentário se for um sistema ou Glean

        grupos_str = ", ".join(grupos) # Converte a lista de grupos em uma string
        conteudo += f" - comentário {idx} ({autor_email} | Grupos: {grupos_str}): {corpo}\n" # Escreve o comentário no texto completo
    
    return conteudo # Retorna o texto completo do ticket
##--------------------------------------------------------------------------##
def get_user_info(user_id):
    url = f"https://{ZENDESK_DOMAIN}.zendesk.com/api/v2/users/{user_id}.json" # URL da API do Zendesk para pegar o usuário
    auth = (ZENDESK_EMAIL, ZENDESK_TOKEN) # Autenticação com email e token
    res = requests.get(url, headers=ZENDESK_HEADERS, auth=auth) # Faz a requisição GET
    if res.status_code != 200: # Verifica se a requisição foi bem sucedida
        return "Erro Desconhecido ao buscar email", [] # Retorna erro se não foi bem sucedida
    user_data = res.json().get("user", {}) # Pega os dados do usuário
    email = user_data.get("email", "Sem email") # Pega o email do usuário
    groups_url = f"https://{ZENDESK_DOMAIN}.zendesk.com/api/v2/users/{user_id}/groups.json"
    response = requests.get(groups_url, auth=auth)
    groups = [] # Inicializa a lista de grupos
    if response.status_code == 200:
        groups = response.json().get("groups", [])
        group_names = [g.get("name", "Sem nome") for g in groups]
        return email, group_names
    else:
        print(f"Erro ao buscar grupos do usuário {user_id}: {response.status_code}")
        return ["Erro ao buscar grupos"]
##--------------------------------------------------------------------------##
def ask_glean(texto_ticket_completo, application_id): # Envia o texto do ticket para a Glean e retorna a resposta
    system_prompt = (
    "Você receberá o conteúdo de um ticket do Zendesk. A estrutura será assim:\n\n"
    "-------------\nTicket ID: <número>\n"
    " - subject: <assunto do ticket>\n"
    " - comentário 1 (<autor> | Grupos: <grupos>): <conteúdo>\n"
    " - comentário 2 (<autor> | Grupos: <grupos>): <conteúdo>\n"
    "...\n\n"
    "Com base nisso, gere uma sugestão de resposta para resolver o problema do cliente de forma clara e útil.\n\n"
    "A resposta deve ser educada e profissional, mantendo um tom amigável.\n\n"
    ) # fim do prompt de sistema
    payload = {
    'stream': True,  # Habilita o streaming
    'applicationId': application_id, # ID do app Product Support
    'messages': list(reversed([
        make_system_message(system_prompt),
        make_content_message(text=texto_ticket_completo)
    ]))
}
    # 📝 Salvar o payload em um arquivo .txt
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"envio_glean_{timestamp}.txt"
    with open(filename, "w", encoding="utf-8") as f:
        f.write("Payload enviado para a Glean:\n\n")
        f.write(json.dumps(payload, indent=2)) 
    response = requests.post(GLEAN_API_URL, headers=GLEAN_HEADERS, json=payload, stream=True) # Envia o payload para a Glean
    #print(response.status_code) # Imprime o status da resposta
    if response.status_code == 200: # Verifica se a resposta foi bem sucedida
        reply,token= process_response_message_stream(response) # Processa a resposta da Glean, ignorando a segunda variável de token
        return reply, token # Retorna a resposta

    else: # Se a resposta não foi bem sucedida
        print("Erro Glean:", response.status_code, response.text) # Imprime o erro
        return None # Retorna None
##--------------------------------------------------------------------------##
def make_system_message(text):  # Novo tipo de mensagem para o sistema da Glean
    return {
        "author": "SYSTEM", 
        "messageType": "CONTENT",
        "fragments": [{"text": text}]
    }
##--------------------------------------------------------------------------##
def make_content_message(author='USER', text=None, citations=None): # Cria uma mensagem de conteúdo no formato de recebimento da API da Glean
    message = { #inicializa a mensagem com autor e o tipo da mensagem
        'author': author, # Define o autor da mensagem
        'messageType': 'CONTENT' # Define o tipo da mensagem como conteúdo
    }
    if text: # Verifica se o texto foi fornecido
        message['fragments'] = [{'text': text}] # Adiciona o texto à mensagem
    if citations: # Verifica se as citações foram fornecidas
        message['citations'] = citations # Adiciona as citações à mensagem
    return message # Retorna a mensagem
##--------------------------------------------------------------------------##
def process_response_message_stream(response): 
    resposta_texto = ''
    todas_citacoes = []
    token = None
    print("Iniciando o processamento da resposta da Glean...")
    for line in response.iter_lines():
        if line:
            line_json = json.loads(line)
           # ✅ Salva o token apenas se ainda não foi capturado
            messages = line_json.get('messages', [])
            for msg in messages:
                if token is None:
                    token = msg.get('messageTrackingToken')
                texto, citacoes = process_message_fragment(msg)
                resposta_texto += texto
                todas_citacoes += citacoes
    # Remover duplicatas com base na URL
    citacoes_unicas = []
    urls_vistas = set()
    print("Removendo duplicatas...")
    for citacao in todas_citacoes:
        url = citacao.get("url") or citacao.get("sourceDocument", {}).get("url")
        if url and url not in urls_vistas:
            urls_vistas.add(url)
            citacoes_unicas.append(citacao)
    # Montar seção de fontes se houver citações únicas
    print("Montando seção de fontes...")
    if citacoes_unicas:
        resposta_texto += "\n\n🔍 *Fontes mencionadas pela Glean:*\n"
        for i, citacao in enumerate(citacoes_unicas, start=1):
            fonte_texto = citacao.get("text", "").strip()

            if not fonte_texto:
                fonte_texto = citacao.get("sourceDocument", {}).get("title", "").strip()

            if not fonte_texto:
                fonte_texto = citacao.get("url", "").strip()

            if not fonte_texto:
                fonte_texto = citacao.get("sourceDocument", {}).get("url", "").strip()

            if not fonte_texto:
                continue  # Ignora citações sem conteúdo útil

            url = citacao.get("url") or citacao.get("sourceDocument", {}).get("url")
            if url:
                resposta_texto += f"{i}. [{fonte_texto}]({url})\n"
            else:
                resposta_texto += f"{i}. {fonte_texto}\n"
    print("Processamento concluído.")
    return resposta_texto, token
##--------------------------------------------------------------------------##
def process_message_fragment(message): # Processa uma mensagem fragmentada e retorna o texto e as citações
    text = '' # Inicializa o texto vazio
    citations = [] # Inicializa as citações vazias
    if message['messageType'] == 'CONTENT': # Verifica se o tipo da mensagem é conteúdo
        for fragment in message.get('fragments', []): # Itera pelos fragmentos da mensagem
            text += fragment.get('text', '') # Adiciona o texto do fragmento ao texto
        for citation in message.get('citations', []): #Itera pelas citaçoes da mensagem
            citations += message.get('citations', []) # Adiciona as citações da mensagem
    return text, citations # Retorna o texto e as citações
##--------------------------------------------------------------------------##
def salvar_resposta_em_txt(ticket_id, resposta):
    with open(f"resposta_ticket_{ticket_id}.txt", "w", encoding="utf-8") as f:
        f.write(resposta)
    print(f"Resposta da Glean salva em resposta_ticket_{ticket_id}.txt")
##--------------------------------------------------------------------------##
def post_internal_note_to_zendesk(ticket_id, note_text):
    url = f"https://{ZENDESK_DOMAIN}.zendesk.com/api/v2/tickets/{ticket_id}.json"
    auth = (ZENDESK_EMAIL, ZENDESK_TOKEN)
    aviso = (
        "⚠️ This is a suggestion made automatically by a *pilot version* of Glean Assistant for Zendesk. I am triggered by tagging Glean on any ticket!\n\n"
        "Please review the veracity and clarity of the answer before sending to the client. Any feedback can be sent to #glean-hub!\n\n"
    )
    payload = {
        "ticket": {
            "comment": {
                "body": aviso + note_text,
                "public": False
            }
        }
    }
    try:
        res = requests.put(url, headers=ZENDESK_HEADERS, auth=auth, json=payload)
        if res.status_code == 200:
            print(f"Resposta postada no ticket {ticket_id}")
        else:
            print(f"Erro ao postar resposta: {res.status_code}, {res.text}")
    except Exception as e:
        print(f"Exceção ao postar resposta no Zendesk: {e}")
##--------------------------------------------------------------------------##
def buscar_formulario_para_tickets(ticket_id):
        # 1. Buscar os dados do ticket
        url = f"https://{ZENDESK_DOMAIN}.zendesk.com/api/v2/tickets/{ticket_id}.json"
        auth = (ZENDESK_EMAIL, ZENDESK_TOKEN) # Autenticação com email e token
        resp = requests.get(url, auth=auth, headers=ZENDESK_HEADERS)
        if resp.status_code == 200:
            ticket = resp.json().get("ticket", {})
            ticket_form_id = ticket.get("ticket_form_id")
            return ticket_form_id
        else:
            print(f"❌ Erro ao buscar ticket ID {ticket_id}: {resp.status_code} - {resp.text}")
            return None
##--------------------------------------------------------------------------##
def processa_ticket(data):
    ticket_id = data.get("ticket", {}).get("id")
    form_id = str(buscar_formulario_para_tickets(ticket_id))
    #print("ticket_id extraído:", ticket_id)
    print("form_id:", form_id, type(form_id))
    print("FSE_ZENDESK_ID:", FSE_ZENDESK_ID, type(FSE_ZENDESK_ID))
    print("PS_ZENDESK_ID:", PS_ZENDESK_ID, type(PS_ZENDESK_ID))
    print("FIN_ZENDESK_ID:", FIN_ZENDESK_ID, type(FIN_ZENDESK_ID))
    #print("PS_ID:", PS_ID)
    #print("FIN_ID:", FIN_ID)
    if not ticket_id:
        return
    if form_id == PS_ZENDESK_ID: # Verifica se o ID do formulário é um dos IDs de Product Support
        print("ENTROU 1")
        application_id = PS_ID
    elif form_id == FSE_ZENDESK_ID: # Verifica se o ID do formulário é um dos IDs de FSE
        print("ENTROU 2")   
        application_id = FSE_ID
    elif form_id == FIN_ZENDESK_ID: # Verifica se o ID do formulário é um dos IDs de Finance
        print("ENTROU 3")
        application_id = FIN_ID    
    else:
        print("ENTROU 4")
        application_id = FSE_ID # Se não for nenhum dos IDs, não faz nada
    print("application_id:", application_id)
    ticket = buscar_dados_completos_do_ticket(ticket_id)
    comentarios = buscar_comentarios_do_ticket(ticket_id)
    texto_ticket_completo = gerar_texto_completo_do_ticket(ticket_id, ticket, comentarios)
    response_from_glean, token = ask_glean(texto_ticket_completo, application_id)

    # 📝 Salvar token e ticket_id no Excel
    if token:
        salvar_token_em_excel(ticket_id, token)
    if response_from_glean:
        post_internal_note_to_zendesk(ticket_id, response_from_glean)
        salvar_resposta_em_txt(ticket_id, response_from_glean)
##--------------------------------------------------------------------------##
def salvar_token_em_excel(ticket_id, token, arquivo="tokens.xlsx"):
    if os.path.exists(arquivo):
        wb = load_workbook(arquivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["ticket_id", "tracking_token"])  # cabeçalho

    ws.append([ticket_id, token])
    wb.save(arquivo)
    print(f"✅ Token salvo para o ticket {ticket_id}: {token}")
##--------------------------------------------------------------------------##
##--------------------------------------------------------------------------##
## Função que recebe o webhook do Zendesk e chama as outras funções
@app.route("/zendesk-to-glean", methods=["POST"]) # Rota do webhook com método de post
def zendesk_webhook():
    data = request.json
    print("Payload recebido:")
    print(json.dumps(data, indent=2))

    # Dispara a thread e responde já
    threading.Thread(target=processa_ticket, args=(data,)).start()
    return {"status": "received"}, 200

##--------------------------------------------------------------------------##
## Função principal para executar o Flask apenas quando o script está sendo chamado diretamente
if __name__ == "__main__": # Executa o Flask
    app.run(port=5001, debug=True) # Define a porta e ativa o modo debug
